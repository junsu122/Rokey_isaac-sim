"""
md 문서들을 Excel 파일로 변환합니다.
출력: docs/rokey_factory_docs.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── 색상 팔레트 ────────────────────────────────────────────────
C = {
    "header_dark":  "1E293B",  # 진한 남색 (시트 헤더)
    "header_mid":   "1E3A5F",  # 중간 파란 (섹션 헤더)
    "header_light": "DBEAFE",  # 연한 파란 (테이블 헤더)
    "row_even":     "F8FAFC",  # 짝수 행
    "row_odd":      "FFFFFF",  # 홀수 행
    "accent_green": "D1FAE5",  # 초록 강조
    "accent_yellow":"FEF9C3",  # 노란 강조
    "accent_red":   "FEE2E2",  # 빨간 강조
    "accent_purple":"EDE9FE",  # 보라 강조
    "text_white":   "FFFFFF",
    "text_dark":    "1E293B",
    "text_blue":    "1D4ED8",
    "border_color": "CBD5E1",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="1E293B", size=11):
    return Font(bold=bold, color=color, size=size, name="맑은 고딕")

def border():
    s = Side(border_style="thin", color=C["border_color"])
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_title(ws, row, text, n_cols=8):
    ws.row_dimensions[row].height = 36
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True, color=C["text_white"], size=14, name="맑은 고딕")
    c.fill = fill(C["header_dark"])
    c.alignment = center()

def write_section(ws, row, text, n_cols=8):
    ws.row_dimensions[row].height = 24
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True, color=C["text_white"], size=11, name="맑은 고딕")
    c.fill = fill(C["header_mid"])
    c.alignment = left()

def write_table_header(ws, row, headers, col_widths=None):
    ws.row_dimensions[row].height = 20
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = font(bold=True, color=C["text_white"])
        c.fill = fill("2563EB")
        c.alignment = center()
        c.border = border()
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def write_row(ws, row, values, row_idx=0, accent=None):
    ws.row_dimensions[row].height = 18
    bg = C["row_even"] if row_idx % 2 == 0 else C["row_odd"]
    if accent:
        bg = accent
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v)
        c.font = font()
        c.fill = fill(bg)
        c.alignment = left()
        c.border = border()

def write_note(ws, row, text, n_cols=8):
    ws.row_dimensions[row].height = 16
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    c = ws.cell(row, 1, text)
    c.font = Font(italic=True, color="64748B", size=10, name="맑은 고딕")
    c.alignment = left()


# ══════════════════════════════════════════════════════════════
# 시트 1: 시스템 아키텍처 개요
# ══════════════════════════════════════════════════════════════
def sheet_architecture(wb):
    ws = wb.create_sheet("시스템 아키텍처")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 20

    r = 1
    write_title(ws, r, "Rokey Factory — 시스템 아키텍처 개요", 5); r += 2

    # 데이터 레이어 흐름
    write_section(ws, r, "데이터 레이어 흐름", 5); r += 1
    write_table_header(ws, r, ["레이어", "구성 요소", "역할", "통신 방식", "비고"], [22,28,30,22,18]); r += 1
    layers = [
        ("시뮬레이션\n(Isaac Sim)", "AMR / 드론 / M0609 암\n카메라 (ArUco 감지)", "로봇 동작 수행\n마커 인식", "ROS2 토픽 발행", "NVIDIA Omniverse"),
        ("ROS2 브릿지\n(run_bridge.py)", "AMRBridge\nDroneBridge\nArmBridge", "ROS2 ↔ Firestore\n양방향 동기화", "Firebase Admin SDK\n(Python)", "update_interval=0.5s"),
        ("클라우드 DB\n(Firebase)", "Firestore\nrokey-factory-base", "실시간 상태 저장\n작업 관리", "REST / WebSocket", "6개 컬렉션"),
        ("모니터링\n대시보드", "React + TypeScript\nVite + Tailwind CSS", "실시간 위치 맵\n로봇 상태 카드", "Firestore onSnapshot", "http://localhost:5173"),
    ]
    for i, row_data in enumerate(layers):
        write_row(ws, r, row_data, i)
        ws.row_dimensions[r].height = 36
        for col in range(1, 6):
            ws.cell(r, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r += 1

    r += 1
    write_section(ws, r, "로봇 구성", 5); r += 1
    write_table_header(ws, r, ["로봇 ID", "종류", "주요 역할", "Firebase 문서", "ROS2 토픽 (주)"]); r += 1
    robots = [
        ("amr_001",  "자율주행 로봇 (AMR)", "물품 운반 / 구획 이동",    "robots/amr_001",   "/amr_001/odom"),
        ("drone_001","드론",                "공중 배송 / 정찰",          "robots/drone_001", "/drone_001/odom"),
        ("m0609",    "두산 협동로봇 암",    "ArUco 인식 / 물품 픽업",    "robots/m0609",     "/m0609/joint_states"),
    ]
    accents = [C["accent_green"], C["accent_yellow"], C["accent_purple"]]
    for i, row_data in enumerate(robots):
        write_row(ws, r, row_data, i, accents[i]); r += 1

    r += 1
    write_section(ws, r, "ArUco 마커 역할 분류", 5); r += 1
    write_table_header(ws, r, ["마커 ID 범위", "역할", "부착 위치", "인식 시 동작", "등록 수"]); r += 1
    markers = [
        ("0 ~ 4",   "item",        "제품 박스",      "tasks 생성, 암 픽업 시작",          "5개"),
        ("10 ~ 14", "section",     "선반/바닥",       "AMR 위치 보정, 도착 확인",          "5개"),
        ("20 ~ 22", "destination", "최종 배송지",     "배송 완료 처리, 작업 complete",     "3개"),
    ]
    accents2 = [C["accent_yellow"], C["accent_green"], C["accent_red"]]
    for i, row_data in enumerate(markers):
        write_row(ws, r, row_data, i, accents2[i]); r += 1

    r += 1
    write_section(ws, r, "물품 → 배송지 매핑", 5); r += 1
    write_table_header(ws, r, ["마커 ID", "물품명", "정렬 구획", "최종 배송지", "배송지 좌표"]); r += 1
    items = [
        ("0", "Apple Watch",  "A-1", "Gangnam (강남)",      "(1.5, 0.5)"),
        ("3", "AirPods",      "A-1", "Gangnam (강남)",      "(1.5, 0.5)"),
        ("1", "Galaxy Tab",   "A-2", "Seocho (서초)",       "(1.5, 0.0)"),
        ("4", "Kindle",       "A-2", "Seocho (서초)",       "(1.5, 0.0)"),
        ("2", "MacBook Pro",  "A-3", "Guro Digital (구로)", "(1.5, -0.5)"),
    ]
    for i, row_data in enumerate(items):
        write_row(ws, r, row_data, i); r += 1


# ══════════════════════════════════════════════════════════════
# 시트 2: Firestore 데이터베이스 구조
# ══════════════════════════════════════════════════════════════
def sheet_database(wb):
    ws = wb.create_sheet("Firestore DB 구조")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [18, 22, 26, 22, 18]):
        ws.column_dimensions[col].width = w

    r = 1
    write_title(ws, r, "Firestore 데이터베이스 구조 — rokey-factory-base", 5); r += 2

    # ── robots/ ──────────────────────────────────────────────
    write_section(ws, r, "컬렉션: robots/  (로봇 실시간 상태)", 5); r += 1
    write_table_header(ws, r, ["문서 ID", "필드명", "타입 / 예시값", "설명", "업데이트 주기"]); r += 1

    amr_fields = [
        ("amr_001", "battery",           "float  예: 85.3",         "배터리 잔량 (%)",            "ROS2 battery_state"),
        ("",        "charge_status",     "string  operating / charging", "충전 상태",             "수동 호출"),
        ("",        "cargo_status",      "string  empty / loading /\ntransporting / unloading", "물품 적재 상태", "작업 시작/완료 시"),
        ("",        "position.x/y/yaw",  "float  예: {x:0.4, y:0.3, yaw:90}", "현재 위치 (오도메트리)", "ROS2 /odom"),
        ("",        "speed",             "float  예: 0.35",         "현재 속도 (m/s)",            "ROS2 /odom"),
        ("",        "current_task",      "string  예: task_4905",   "담당 중인 작업 ID",          "작업 할당 시"),
        ("",        "localization",      "map  {marker_id, label,\nestimated_pos, distance}", "ArUco 위치 보정 결과", "섹션 마커 인식 시"),
        ("drone_001","battery",          "float",                   "배터리 잔량",                "ROS2"),
        ("",        "position.x/y/z",   "float  예: {x:0.0, y:0.0, z:2.5}", "3D 위치",         "ROS2 /odom"),
        ("",        "altitude",          "float  예: 2.5",          "현재 고도 (m)",              "ROS2"),
        ("",        "heading",           "float  0~360",            "방향 (0=북)",                "ROS2"),
        ("",        "cargo_status",      "string  (AMR와 동일)",    "물품 적재 상태",             "작업 시"),
        ("m0609",  "status",            "string  idle / picking /\nplacing / moving / error", "로봇 동작 상태", "동작 시작/완료"),
        ("",        "gripper",           "string  open / closed",   "그리퍼 상태",               "픽업 시"),
        ("",        "position.x/y/z",   "float",                   "엔드이펙터 위치",            "ROS2 end_effector_pose"),
        ("",        "joints",            "array[6]  float",         "6축 관절값 (degrees)",       "ROS2 joint_states"),
        ("",        "detected_item",     "map  {marker_id, label,\ncategory, position_xyz}", "최근 인식 물품", "ArUco 인식 시"),
    ]
    for i, row_data in enumerate(amr_fields):
        write_row(ws, r, row_data, i); r += 1

    r += 1
    write_section(ws, r, "컬렉션: sections/  (창고 구획 마스터)", 5); r += 1
    write_table_header(ws, r, ["section_id", "위치 (x, y)", "ArUco 마커 ID", "capacity", "비고"]); r += 1
    sections = [
        ("A-1", "(-0.4,  0.3)", "10", "5", "Gangnam 배송 물품"),
        ("A-2", "( 0.0,  0.3)", "11", "5", "Seocho 배송 물품"),
        ("A-3", "( 0.4,  0.3)", "12", "5", "Guro 배송 물품"),
        ("B-1", "(-0.4, -0.3)", "13", "5", "미사용 (확장 예약)"),
        ("B-2", "( 0.0, -0.3)", "14", "5", "미사용 (확장 예약)"),
    ]
    for i, row_data in enumerate(sections):
        write_row(ws, r, row_data, i); r += 1

    r += 1
    write_section(ws, r, "컬렉션: tasks/  (로봇 작업 지시서)", 5); r += 1
    write_table_header(ws, r, ["필드명", "타입 / 예시값", "설명", "M0609 작업", "AMR 작업"]); r += 1
    tasks = [
        ("task_id",      "string  task_49051116",  "자동 생성 UUID 8자",   "task_49051116",  "task_80d3b2d7"),
        ("item_id",      "string  ITEM-2836C439",  "items/ 문서 참조",     "ITEM-2836C439",  "ITEM-2836C439"),
        ("robot_id",     "string",                 "담당 로봇 ID",          "m0609",          "amr_001"),
        ("destination",  "string",                 "이동 목표",             "A-1 (정렬 구획)","Gangnam (배송지)"),
        ("status",       "string  pending /\nin_progress / completed / failed", "작업 상태", "pending→completed", "pending→completed"),
        ("created_at",   "timestamp",              "작업 생성 시각",        "마커 인식 시",   "마커 인식 시"),
        ("started_at",   "timestamp / null",       "시작 시각",             "",               ""),
        ("completed_at", "timestamp / null",       "완료 시각",             "",               ""),
    ]
    for i, row_data in enumerate(tasks):
        write_row(ws, r, row_data, i); r += 1

    r += 1
    write_section(ws, r, "컬렉션: navigation/  (AMR 이동 명령)", 5); r += 1
    write_table_header(ws, r, ["필드명", "타입 / 예시값", "설명", "status 값", "의미"]); r += 1
    navs = [
        ("current_target",   "string  A-1",                 "현재 이동 목표",      "idle",       "대기 중"),
        ("target_position",  "map  {x:-0.4, y:0.3}",        "목표 좌표",           "navigating", "이동 중"),
        ("assigned_item_id", "string  ITEM-xxxx",           "운반 중인 물품 ID",   "arrived",    "도착 확인됨"),
        ("status",           "string  navigating / arrived","네비게이션 상태",     "",           ""),
        ("confirmed_section","string / null",                "마커로 확인된 위치",  "",           ""),
    ]
    for i, row_data in enumerate(navs):
        write_row(ws, r, row_data, i); r += 1


# ══════════════════════════════════════════════════════════════
# 시트 3: ROS2 토픽 목록
# ══════════════════════════════════════════════════════════════
def sheet_ros_topics(wb):
    ws = wb.create_sheet("ROS2 토픽")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [30, 28, 22, 22, 20]):
        ws.column_dimensions[col].width = w

    r = 1
    write_title(ws, r, "ROS2 토픽 목록 — Isaac Sim ↔ Firebase 브릿지", 5); r += 2

    write_section(ws, r, "Isaac Sim → Firebase  (ROS2 구독 토픽)", 5); r += 1
    write_table_header(ws, r, ["토픽 이름", "메시지 타입", "담당 브릿지", "Firebase 저장 위치", "비고"]); r += 1
    sub_topics = [
        ("/amr_001/odom",             "nav_msgs/Odometry",          "AMRBridge",   "robots/amr_001\n.position, speed", "update_interval 적용"),
        ("/amr_001/battery_state",    "sensor_msgs/BatteryState",   "AMRBridge",   "robots/amr_001\n.battery",         ""),
        ("/amr_001/cmd_vel",          "geometry_msgs/Twist",        "AMRBridge",   "로그만 (저장 없음)",                "빈번 발행"),
        ("/drone_001/odom",           "nav_msgs/Odometry",          "DroneBridge", "robots/drone_001\n.position, altitude", ""),
        ("/drone_001/battery_state",  "sensor_msgs/BatteryState",   "DroneBridge", "robots/drone_001\n.battery",       ""),
        ("/m0609/joint_states",       "sensor_msgs/JointState",     "ArmBridge",   "robots/m0609\n.joints",            ""),
        ("/m0609/end_effector_pose",  "geometry_msgs/PoseStamped",  "ArmBridge",   "robots/m0609\n.position",          ""),
        ("/m0609/task_done",          "std_msgs/String (JSON)",     "ArmBridge",   "tasks/ status=completed",          "{task_id, result}"),
    ]
    for i, row_data in enumerate(sub_topics):
        write_row(ws, r, row_data, i)
        ws.row_dimensions[r].height = 30
        for col in range(1, 6):
            ws.cell(r, col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 1
    write_section(ws, r, "Firebase → Isaac Sim  (ROS2 발행 토픽)", 5); r += 1
    write_table_header(ws, r, ["토픽 이름", "메시지 타입", "담당 브릿지", "트리거 조건", "메시지 형식"]); r += 1
    pub_topics = [
        ("/amr_001/goal",             "geometry_msgs/PoseStamped",  "AMRBridge",   "navigation/amr_001\n.status = navigating",   "PoseStamped(x,y)"),
        ("/amr_001/firebase_status",  "std_msgs/String (JSON)",     "AMRBridge",   "navigation 문서 변경 시",                    '{status, target}'),
        ("/drone_001/pose_command",   "geometry_msgs/PoseStamped",  "DroneBridge", "드론 이동 명령 시",                          "PoseStamped(x,y,z)"),
        ("/drone_001/firebase_status","std_msgs/String (JSON)",     "DroneBridge", "로봇 상태 변경 시",                          '{status}'),
        ("/m0609/task_command",       "std_msgs/String (JSON)",     "ArmBridge",   "tasks/ 에 pending 작업\n생성 시",            '{task_id, action, destination}'),
        ("/m0609/firebase_status",    "std_msgs/String (JSON)",     "ArmBridge",   "로봇 상태 변경 시",                          '{status}'),
        ("/aruco/detections",         "std_msgs/String (JSON)",     "ArucoBridge", "ArUco 마커 인식 시",                         '{marker_id, role, label}'),
    ]
    for i, row_data in enumerate(pub_topics):
        write_row(ws, r, row_data, i, accent=C["accent_green"])
        ws.row_dimensions[r].height = 30
        for col in range(1, 6):
            ws.cell(r, col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 1
    write_section(ws, r, "Isaac Sim OmniGraph 설정 요약", 5); r += 1
    write_table_header(ws, r, ["로봇", "OmniGraph 노드", "방향", "토픽", "targetPrim"]); r += 1
    omni = [
        ("AMR",    "ROS2 Publish Odometry",      "→ Firebase", "/amr_001/odom",           "/World/amr_001"),
        ("AMR",    "ROS2 Subscribe PoseStamped", "← Firebase", "/amr_001/goal",           "/World/amr_001"),
        ("드론",   "ROS2 Publish Odometry",      "→ Firebase", "/drone_001/odom",          "/World/drone_001"),
        ("드론",   "ROS2 Subscribe PoseStamped", "← Firebase", "/drone_001/pose_command",  "/World/drone_001"),
        ("M0609",  "ROS2 Publish JointState",    "→ Firebase", "/m0609/joint_states",      "/World/m0609"),
        ("M0609",  "ROS2 Subscribe String",      "← Firebase", "/m0609/task_command",      "/World/m0609"),
        ("M0609",  "ROS2 Publish String",        "→ Firebase", "/m0609/task_done",         "/World/m0609"),
    ]
    for i, row_data in enumerate(omni):
        write_row(ws, r, row_data, i); r += 1


# ══════════════════════════════════════════════════════════════
# 시트 4: 배송 시나리오 흐름
# ══════════════════════════════════════════════════════════════
def sheet_scenario(wb):
    ws = wb.create_sheet("배송 시나리오")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [6, 26, 22, 28, 24, 20]):
        ws.column_dimensions[col].width = w

    r = 1
    write_title(ws, r, "물품 인식 → 배송 완료 시나리오 흐름", 6); r += 2

    write_section(ws, r, "예시: Apple Watch (마커 ID=0) 인식 → 강남 배송", 6); r += 1
    write_table_header(ws, r, ["단계", "이벤트", "담당", "Firebase 변경", "ROS2 동작", "비고"]); r += 1

    steps = [
        ("1", "카메라가 ID=0\n(Apple Watch) 인식", "robot_main.py",
         "items/ 생성\ntasks/ ×2 생성",
         "없음", "section=A-1\ndestination=Gangnam"),
        ("2", "M0609 픽업 명령", "ArmBridge",
         "robots/m0609\n.status=picking",
         "/m0609/task_command\n발행", "action=pick"),
        ("3", "AMR 이동 명령", "AMRBridge",
         "navigation/amr_001\n.status=navigating\n.current_target=A-1",
         "/amr_001/goal 발행\n(x=-0.4, y=0.3)", ""),
        ("4", "AMR 이동 중\n/amr_001/odom 수신", "AMRBridge",
         "robots/amr_001\n.position 업데이트\n(0.5s 간격)",
         "오도메트리 구독", "대시보드 맵 실시간 갱신"),
        ("5", "A-1 섹션 마커(ID=10) 인식", "robot_main.py",
         "robots/amr_001\n.localization 업데이트\nnavigation/ status=arrived",
         "없음", "위치 보정"),
        ("6", "물품 하역 완료", "AMRBridge",
         "robots/amr_001\n.cargo_status=unloading\ntasks/m0609 completed",
         "/m0609/task_done 수신", ""),
        ("7", "AMR → 배송지(강남) 이동", "AMRBridge",
         "navigation/amr_001\n.current_target=Gangnam",
         "/amr_001/goal 발행\n(x=1.5, y=0.5)", ""),
        ("8", "강남 마커(ID=20) 인식\n배송 완료", "robot_main.py",
         "items/ status=delivered\ntasks/amr completed\nnavigation/ status=idle",
         "없음", "전체 작업 종료"),
    ]

    step_accents = [
        C["row_odd"], C["accent_purple"], C["accent_green"], C["row_odd"],
        C["accent_yellow"], C["accent_green"], C["row_odd"], C["accent_red"],
    ]
    for i, (step_data, accent) in enumerate(zip(steps, step_accents)):
        write_row(ws, r, step_data, i, accent)
        ws.row_dimensions[r].height = 40
        for col in range(1, 7):
            ws.cell(r, col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    r += 1
    write_section(ws, r, "실행 명령 모음", 6); r += 1
    write_table_header(ws, r, ["목적", "명령어", "설명", "실행 위치", "선행 조건", "비고"]); r += 1
    cmds = [
        ("재고 초기화",    "python3 DB/reset_inventory.py",          "Firestore 전체 초기화",          "robot/", "Firebase 연결",    "최초 1회"),
        ("재고 등록",      "python3 DB/setup_inventory.py",          "yaml 기반 초기 데이터 등록",     "robot/", "reset 후",         "최초 1회"),
        ("브릿지 전체",    "python3 ros_bridge/run_bridge.py",             "3개 브릿지 동시 실행",           "robot/", "Isaac Sim 실행 후","ROS2 환경 필요"),
        ("AMR만",          "python3 ros_bridge/run_bridge.py --amr-only",  "AMR 브릿지만 실행",              "robot/", "ROS2",             ""),
        ("웹캠 테스트",    "python3 robot_main.py --webcam 0",       "Isaac Sim 없이 웹캠 테스트",     "robot/", "없음",             ""),
        ("Firestore 모니터","python3 DB/monitor.py --watch",         "터미널 실시간 모니터링",         "robot/", "Firebase 연결",    ""),
        ("대시보드 실행",  "cd UI && npm run dev",        "웹 대시보드 시작",               "code/",        ".env.local 설정",  "localhost:5173"),
    ]
    for i, row_data in enumerate(cmds):
        write_row(ws, r, row_data, i)
        ws.row_dimensions[r].height = 24
        for col in range(1, 7):
            ws.cell(r, col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1


# ══════════════════════════════════════════════════════════════
# 시트 5: 위치 좌표 설정
# ══════════════════════════════════════════════════════════════
def sheet_coordinates(wb):
    ws = wb.create_sheet("위치 좌표 설정")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [18, 14, 14, 20, 30]):
        ws.column_dimensions[col].width = w

    r = 1
    write_title(ws, r, "창고 위치 좌표 설정 (환경 구성 후 수정)", 5); r += 2

    write_section(ws, r, "섹션 좌표  (object_registry.yaml + WarehouseMap.tsx)", 5); r += 1
    write_table_header(ws, r, ["section_id", "x 좌표", "y 좌표", "ArUco 마커 ID", "수정 위치"]); r += 1
    secs = [
        ("A-1", "-0.4", "0.3",  "10", "yaml + WarehouseMap.tsx SECTIONS"),
        ("A-2", " 0.0", "0.3",  "11", "yaml + WarehouseMap.tsx SECTIONS"),
        ("A-3", " 0.4", "0.3",  "12", "yaml + WarehouseMap.tsx SECTIONS"),
        ("B-1", "-0.4", "-0.3", "13", "yaml + WarehouseMap.tsx SECTIONS"),
        ("B-2", " 0.0", "-0.3", "14", "yaml + WarehouseMap.tsx SECTIONS"),
    ]
    for i, row_data in enumerate(secs):
        write_row(ws, r, row_data, i); r += 1

    r += 1
    write_section(ws, r, "배송지 좌표  (WarehouseMap.tsx DESTINATIONS)", 5); r += 1
    write_table_header(ws, r, ["배송지", "x 좌표", "y 좌표", "ArUco 마커 ID", "수정 위치"]); r += 1
    dests = [
        ("Gangnam (강남)",      "1.5", " 0.5",  "20", "WarehouseMap.tsx DESTINATIONS"),
        ("Seocho (서초)",       "1.5", " 0.0",  "21", "WarehouseMap.tsx DESTINATIONS"),
        ("Guro Digital (구로)", "1.5", "-0.5",  "22", "WarehouseMap.tsx DESTINATIONS"),
    ]
    for i, row_data in enumerate(dests):
        write_row(ws, r, row_data, i, accent=C["accent_yellow"]); r += 1

    r += 1
    write_section(ws, r, "로봇 암 고정 위치  (WarehouseMap.tsx ARM_WORLD)", 5); r += 1
    write_table_header(ws, r, ["항목", "현재값", "단위", "수정 파일", "수정 상수"]); r += 1
    arm_pos = [
        ("x 좌표", "1.0", "m", "WarehouseMap.tsx", "ARM_WORLD.x"),
        ("y 좌표", "0.0", "m", "WarehouseMap.tsx", "ARM_WORLD.y"),
    ]
    for i, row_data in enumerate(arm_pos):
        write_row(ws, r, row_data, i, accent=C["accent_purple"]); r += 1

    r += 1
    write_note(ws, r, "※ USD 씬 파일(Isaac Sim)을 기반으로 실제 환경 구성 완료 후 위 좌표를 업데이트하세요.", 5); r += 1
    write_note(ws, r, "※ object_registry.yaml 수정 시 Firestore setup_inventory.py 재실행 필요합니다.", 5)


# ══════════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    wb.remove(wb.active)  # 기본 Sheet 제거

    sheet_architecture(wb)
    sheet_database(wb)
    sheet_ros_topics(wb)
    sheet_scenario(wb)
    sheet_coordinates(wb)

    out = "docs/rokey_factory_docs.xlsx"
    wb.save(out)
    print(f"✓ 저장 완료: {out}")
    print(f"  시트 목록: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
